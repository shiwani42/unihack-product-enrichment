#!/usr/bin/env python3
"""Import UniHack dashboard reference files into machine-readable lookups.

Drop the organizer files into guidelines/references/ (exact names below) and
run this script. Each import is optional: whatever exists gets converted to
data/reference/*.json which the rest of the pipeline consumes automatically.

Expected inputs (guidelines/references/):
  Unicat_Lov_v1_0_Updated_With_Remarks.xlsx      -> lov_values.json (~161k LOV rows)
  Unilog_Master_UOM_Standards_Abbreviations_and_Terms.xlsx -> uom_standards.json
  UniCat_Manufacturer_and_Brand_List.xlsx        -> manufacturers.json (27k rows)
  Decimal_Fraction.xlsx                          -> fraction_inch.json (63 conversions)
  Unilog-Sample_200_Items-Input-vs-Output.xlsx   -> reference200_input.csv + reference200_expected.csv

Usage:
  PYTHONPATH=. python3 scripts/import_references.py [--src guidelines/references]
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import threading
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

REF_DIR = ROOT / "guidelines" / "references"
OUT_DIR = ROOT / "data" / "reference"
REFERENCE200_DIR = ROOT / "data" / "reference200"
OFFICIAL_LEAVES_PATH = ROOT / "data" / "taxonomy" / "official_leaves.json"
RUNTIME_REF_DIR = Path("/tmp/unilog/reference")

PLACEHOLDER_TOKENS = {"-- unbranded --", "-- no unilog brand --", "-- no dib brand --"}

_ENSURE_LOCK = threading.Lock()
_ENSURED = False

_LOV_EXACT = "Unicat_Lov_v1_0_Updated_With_Remarks.xlsx"
_UOM_EXACT = "Unilog_Master_UOM_Standards_Abbreviations_and_Terms.xlsx"
_MFR_EXACT = "UniCat_Manufacturer_and_Brand_List.xlsx"
_FRAC_EXACT = "Decimal_Fraction.xlsx"
_REF200_EXACT = "Unilog-Sample_200_Items-Input-vs-Output.xlsx"


_KIND_CACHE: dict[str, str | None] = {}


def _stem(path: Path) -> str:
    return path.stem.lower().replace(" ", "_").replace("-", "_")


def _peek_workbook(path: Path, max_sheets: int = 4) -> tuple[list[str], list[dict[str, int]]]:
    """Read sheet names and the first header rows only — do not load a 161k-row LOV."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        names = list(workbook.sheetnames)
        maps: list[dict[str, int]] = []
        for sheet_name in names[:max_sheets]:
            rows: list[list[str]] = []
            for index, row in enumerate(workbook[sheet_name].iter_rows(values_only=True)):
                rows.append([_clean(cell) for cell in row])
                if index >= 11:
                    break
            maps.append(_header_map(rows) if rows else {})
        return names, maps
    finally:
        workbook.close()


def classify_workbook(path: Path) -> str | None:
    """Kind from column headers (and sheet names), not the filename."""
    if not path.exists():
        return None
    cache_key = f"{path.resolve()}:{path.stat().st_mtime_ns}"
    if cache_key in _KIND_CACHE:
        return _KIND_CACHE[cache_key]
    try:
        names, maps = _peek_workbook(path)
    except Exception:
        _KIND_CACHE[cache_key] = None
        return None
    kind = _kind_from_headers(names, maps)
    _KIND_CACHE[cache_key] = kind
    return kind


def _kind_from_headers(sheet_names: list[str], maps: list[dict[str, int]]) -> str | None:
    lowered_names = " ".join(name.lower() for name in sheet_names)
    all_keys = set()
    for header in maps:
        all_keys.update(header)

    def has(*needles: str) -> bool:
        return any(any(needle in key for needle in needles) for key in all_keys)

    lov = has("attribute label", "normalized label") and has("attribute value", "normalized value")
    if not lov:
        lov = "attribute label" in all_keys and ("attribute values" in all_keys or "values" in all_keys)
    if lov:
        return "lov"
    if has("manufacturer_name") and has("brand_name"):
        return "manufacturers"
    if has("abbreviation", "capture form") and has("measurement", "uom type", "uom"):
        return "uom"
    if has("decimal") and has("fraction"):
        return "fractions"
    if has("classpath", "leaf node classpath") and not has("attribute value", "normalized value"):
        return "taxonomy"
    if any(token in lowered_names for token in ("delivery", "expected", "output")) and has("mfg_part_num", "manufacturer_part_number"):
        return "reference200"
    return None


def find_workbook(src: Path, exact: str, *token_groups: tuple[str, ...], kind: str | None = None) -> Path | None:
    """Exact organizer name, then filename tokens, then header-column kind."""
    if not src.exists():
        return None
    exact_path = src / exact
    if exact_path.exists():
        return exact_path
    candidates = [p for p in sorted(src.glob("*.xlsx")) + sorted(src.glob("*.xls")) if not _stem(p).startswith("~$")]
    for path in candidates:
        stem = _stem(path)
        for group in token_groups:
            if group and all(token in stem for token in group):
                return path
    if kind:
        for path in candidates:
            if classify_workbook(path) == kind:
                return path
    return None


def extra_lov_workbooks(src: Path, primary: Path | None) -> list[Path]:
    if not src.exists():
        return []
    found: list[Path] = []
    for path in sorted(src.glob("*.xlsx")) + sorted(src.glob("*.xls")):
        if primary and path.resolve() == primary.resolve():
            continue
        stem = _stem(path)
        if "lov" in stem or "list_of_values" in stem or classify_workbook(path) == "lov":
            found.append(path)
    return found


def reference_search_dirs() -> list[Path]:
    env = (os.environ.get("UNILOG_REFERENCES_DIR") or "").strip()
    ordered = []
    if env:
        ordered.append(Path(env))
    ordered.extend(
        [
            REF_DIR,
            ROOT / "guidelines",
            Path.cwd() / "references",
            Path("/tmp/unilog/references"),
        ]
    )
    seen: set[str] = set()
    dirs: list[Path] = []
    for path in ordered:
        key = str(path.resolve()) if path.exists() else str(path)
        if key in seen:
            continue
        seen.add(key)
        dirs.append(path)
    return dirs


def _writable_out() -> Path:
    try:
        OUT_DIR.mkdir(parents=True, exist_ok=True)
        probe = OUT_DIR / ".write_probe"
        probe.write_text("ok", encoding="utf-8")
        probe.unlink(missing_ok=True)
        return OUT_DIR
    except OSError:
        RUNTIME_REF_DIR.mkdir(parents=True, exist_ok=True)
        return RUNTIME_REF_DIR


def _should_replace(json_path: Path, xlsx: Path | None) -> bool:
    if xlsx is None or not xlsx.exists():
        return False
    if not json_path.exists():
        return True
    try:
        payload = json.loads(json_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError, UnicodeDecodeError):
        return True
    if payload.get("origin") == "mined_sample":
        return True
    try:
        return xlsx.stat().st_mtime > json_path.stat().st_mtime
    except OSError:
        return True


def refresh_reference_consumers(out_dir: Path | None = None) -> None:
    """Point live lookups at freshly imported JSON (including /tmp on Vercel)."""
    directory = out_dir or OUT_DIR
    import identity.brand_resolver as brand_resolver
    import normalize.units as units
    import validate.rules as rules

    rules.REFERENCE_LOV_PATH = directory / "lov_values.json"
    rules._reference_values_cache = None
    brand_resolver.REFERENCE_MANUFACTURERS_PATH = directory / "manufacturers.json"
    brand_resolver._reference_index_cache = None
    units.FRACTION_TABLE = directory / "fraction_inch.json"
    units._fraction_table.cache_clear()
    try:
        from classify.taxonomy_matcher import reset_leaf_cache

        reset_leaf_cache()
    except ImportError:
        pass


def _clean(value) -> str:
    text = "" if value is None else str(value).strip()
    return "" if text.lower() in PLACEHOLDER_TOKENS else text


def _rows(path: Path, sheet_index: int | None = None) -> list[list[str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[sheet_index]] if sheet_index is not None else workbook.active
        return [[_clean(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _header_row_index(rows: list[list[str]]) -> int:
    """Find the header row even with multi-row junk headers (guide warns about this)."""
    best_row, best_score = 0, -1
    for index, row in enumerate(rows[:12]):
        filled = [cell for cell in row if cell]
        if len(filled) >= 3 and len(filled) > best_score:
            best_row, best_score = index, len(filled)
    return best_row


def _header_map(rows: list[list[str]]) -> dict[str, int]:
    return {
        str(cell).strip().lower(): position
        for position, cell in enumerate(rows[_header_row_index(rows)])
        if cell
    }


def _parse_lov_workbook(path: Path) -> dict[str, list[str]] | None:
    rows = _rows(path)
    header = _header_map(rows)

    def col(*candidates: str) -> int | None:
        for name in candidates:
            if name in header:
                return header[name]
        return None

    label_col = col("attribute label", "normalized label", "label")
    value_col = col("attribute values", "normalized values", "value", "values")
    if label_col is None or value_col is None:
        return None
    values_by_label: dict[str, list[str]] = {}
    header_index = _header_row_index(rows)
    for position, row in enumerate(rows):
        if position <= header_index:
            continue
        label = row[label_col] if label_col < len(row) else ""
        value = row[value_col] if value_col < len(row) else ""
        if not label or not value:
            continue
        bucket = values_by_label.setdefault(label.strip(), [])
        if value.strip() not in bucket:
            bucket.append(value.strip())
    return values_by_label


def _merge_lov(into: dict[str, list[str]], extra: dict[str, list[str]]) -> None:
    for label, values in extra.items():
        bucket = into.setdefault(label, [])
        for value in values:
            if value not in bucket:
                bucket.append(value)


def import_lov(src: Path) -> Path | None:
    path = find_workbook(src, _LOV_EXACT, ("unicat", "lov"), ("lov", "remark"), ("list", "of", "values"), kind="lov")
    extras = extra_lov_workbooks(src, path)
    values_by_label: dict[str, list[str]] = {}
    sources: list[str] = []
    if path:
        parsed = _parse_lov_workbook(path)
        if parsed is None:
            print(f"LOV: could not locate label/value columns in {path.name}")
        else:
            values_by_label = parsed
            sources.append(path.name)
    for extra in extras:
        parsed = _parse_lov_workbook(extra)
        if not parsed:
            continue
        _merge_lov(values_by_label, parsed)
        sources.append(extra.name)
    if not values_by_label:
        return None
    payload = {
        "source_file": sources[0] if sources else "",
        "source_files": sources,
        "row_count": sum(len(v) for v in values_by_label.values()),
        "labels": len(values_by_label),
        "values_by_label": values_by_label,
    }
    out = _writable_out() / "lov_values.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, indent=1))
    print(f"LOV: {payload['labels']} labels, {payload['row_count']} values -> {out.name}")
    return out


def import_uom(src: Path) -> Path | None:
    path = find_workbook(src, _UOM_EXACT, ("uom", "standard"), ("abbreviation", "term"), ("master", "uom"), kind="uom")
    if not path:
        return None
    abbreviations: dict[str, str] = {}
    rules: list[str] = []
    for sheet_index in range(3):
        try:
            rows = _rows(path, sheet_index=sheet_index)
        except Exception:
            break
        if not rows:
            continue
        header = _header_map(rows)
        meas_col = next((header[k] for k in header if "measurement" in k or "uom type" in k), None)
        abbr_col = next((header[k] for k in header if "abbreviation" in k or "capture form" in k or "approved" in k), None)
        for row in rows[1:]:
            if meas_col is not None and abbr_col is not None and meas_col < len(row) and abbr_col < len(row):
                measurement, abbr = row[meas_col], row[abbr_col]
                if measurement and abbr:
                    abbreviations.setdefault(measurement.lower(), abbr)
                elif abbr and re.fullmatch(r"[A-Za-z/°%\"']{1,6}", abbr) and abbr not in abbreviations.values():
                    pass
        for row in rows:
            joined = " ".join(cell for cell in row if cell)
            if "space between" in joined.lower() or "hyphen" in joined.lower():
                if joined not in rules:
                    rules.append(joined[:300])
    payload = {
        "source_file": path.name,
        "measurements": len(abbreviations),
        "abbreviations": abbreviations,
        "house_style_rules": rules[:40],
    }
    out = _writable_out() / "uom_standards.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, indent=1))
    print(f"UOM: {payload['measurements']} measurements -> {out.name}")
    return out


def import_manufacturers(src: Path) -> Path | None:
    path = find_workbook(src, _MFR_EXACT, ("manufacturer", "brand"), ("unicat", "brand"), kind="manufacturers")
    if not path:
        return None
    rows = _rows(path)
    header = _header_map(rows)
    mfr_col = next((header[k] for k in header if k.startswith("manufacturer_name")), None)
    brand_col = next((header[k] for k in header if k.startswith("brand_name")), None)
    mfr_code_col = next((header[k] for k in header if k.startswith("manufacturer_code")), None)
    if mfr_col is None:
        print(f"Manufacturers: header columns not found; got {list(header)[:8]}")
        return None
    entries: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for row in rows[1:]:
        mfr = row[mfr_col] if mfr_col < len(row) else ""
        brand = row[brand_col] if brand_col is not None and brand_col < len(row) else ""
        key = (mfr, brand)
        if not mfr or key in seen:
            continue
        seen.add(key)
        entry = {"manufacturer_name": mfr}
        if brand:
            entry["brand_name"] = brand
        if mfr_code_col is not None and mfr_code_col < len(row) and row[mfr_code_col]:
            entry["manufacturer_code"] = row[mfr_code_col]
        entries.append(entry)
    payload = {"source_file": path.name, "count": len(entries), "entries": entries}
    out = _writable_out() / "manufacturers.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, indent=1))
    print(f"Manufacturers: {len(entries)} unique entries -> {out.name}")
    return out


def import_fractions(src: Path) -> Path | None:
    path = find_workbook(src, _FRAC_EXACT, ("decimal", "fraction"), ("fraction", "inch"), kind="fractions")
    if not path:
        return None
    rows = _rows(path)
    mapping: dict[str, str] = {}
    frac_re = re.compile(r"^\d+-?\d*/\d+$|^\d+/\d+$")
    dec_re = re.compile(r"^\d*\.\d+$")
    for row in rows:
        cells = [c for c in row if c]
        for i, cell in enumerate(cells):
            nxt = cells[i + 1] if i + 1 < len(cells) else ""
            if dec_re.fullmatch(cell) and frac_re.fullmatch(nxt):
                decimal = f"{float(cell):.6f}".rstrip("0").rstrip(".")
                mapping.setdefault(decimal, nxt)
            elif frac_re.fullmatch(cell) and dec_re.fullmatch(nxt):
                decimal = f"{float(nxt):.6f}".rstrip("0").rstrip(".")
                mapping.setdefault(decimal, cell)
    payload = {"source_file": path.name, "conversions": len(mapping), "decimal_to_fraction": mapping}
    out = _writable_out() / "fraction_inch.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, indent=1))
    print(f"Fractions: {len(mapping)} conversions -> {out.name}")
    return out


def _sheet_to_csv(xlsx: Path, sheet_name_hint: list[str], out_path: Path) -> bool:
    from openpyxl import load_workbook

    workbook = load_workbook(xlsx, read_only=True, data_only=True)
    try:
        target = None
        for name in workbook.sheetnames:
            lowered = name.lower()
            if any(hint in lowered for hint in sheet_name_hint):
                target = workbook[name]
                break
        target = target or workbook[workbook.sheetnames[0]]
        import csv as csv_mod

        with out_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv_mod.writer(handle)
            for row in target.iter_rows(values_only=True):
                writer.writerow([_clean(cell) for cell in row])
        return True
    finally:
        workbook.close()


def import_reference200(src: Path) -> Path | None:
    path = find_workbook(src, _REF200_EXACT, ("200", "input"), ("sample", "200"), ("input", "output"), kind="reference200")
    if not path:
        return None
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    names = [n.lower() for n in workbook.sheetnames]
    if not names:
        print(f"Reference200: {path.name} has no sheets")
        return None
    workbook.close()

    data_dir = REFERENCE200_DIR
    try:
        data_dir.mkdir(parents=True, exist_ok=True)
    except OSError:
        data_dir = Path("/tmp/unilog/reference200")
        data_dir.mkdir(parents=True, exist_ok=True)
    _sheet_to_csv(path, ["input"], data_dir / "input.csv")
    _sheet_to_csv(path, ["delivery", "output"], data_dir / "expected.csv")
    print(f"Reference200: wrote {data_dir/'input.csv'} and {data_dir/'expected.csv'}")
    return data_dir


_TEMPLATE_HINTS = (
    ("dishwasher", "built_in_dishwasher"),
    ("range", "cooking_range"),
    ("oven", "cooking_range"),
    ("fan", "ceiling_fan"),
    ("deck", "deck_composite"),
    ("coupling", "pipe_fitting"),
    ("fitting", "pipe_fitting"),
    ("faucet", "pipe_fitting"),
    ("wire", "wire_cable"),
    ("cable", "wire_cable"),
    ("grind", "grinding_wheel"),
    ("sand", "sanding_abrasive"),
    ("cutoff", "metal_cutoff_disc"),
    ("cut-off", "metal_cutoff_disc"),
    ("led", "led_lighting"),
    ("lighting", "led_lighting"),
    ("box", "electrical_box"),
    ("trim", "building_trim"),
    ("tool", "power_tool_accessory"),
)


def _template_id_for(classpath: str, fine: str) -> str:
    text = f"{classpath} {fine}".lower()
    for needle, tid in _TEMPLATE_HINTS:
        if needle in text:
            return tid
    return "generic_industrial"


def import_taxonomy(src: Path) -> Path | None:
    """Leaf taxonomy workbook (not the LOV). Keywords come from the leaf name."""
    path = find_workbook(
        src,
        "Unicat_Taxonomy.xlsx",
        ("taxonomy",),
        ("unicat", "categor"),
        ("leaf", "categor"),
        kind="taxonomy",
    )
    if not path or "lov" in _stem(path):
        return None
    rows = _rows(path)
    header = _header_map(rows)

    def col(*candidates: str) -> int | None:
        for name in candidates:
            if name in header:
                return header[name]
        return None

    classpath_col = col("classpath", "leaf node classpath", "leaf classpath")
    if classpath_col is None:
        return None
    if col("attribute values", "normalized values", "attribute label"):
        return None
    dept_col = col("dept", "department")
    class_col = col("class")
    fine_col = col("fine", "leaf", "product name", "item type")
    leaves: list[dict] = []
    seen: set[str] = set()
    header_index = _header_row_index(rows)
    for position, row in enumerate(rows):
        if position <= header_index:
            continue
        classpath = row[classpath_col] if classpath_col < len(row) else ""
        if not classpath or classpath in seen:
            continue
        parts = [part.strip() for part in classpath.split(">") if part.strip()]
        if len(parts) < 2:
            continue
        seen.add(classpath)
        fine = (row[fine_col] if fine_col is not None and fine_col < len(row) else "") or parts[-1]
        dept = (row[dept_col] if dept_col is not None and dept_col < len(row) else "") or parts[0]
        class_name = (row[class_col] if class_col is not None and class_col < len(row) else "") or (
            parts[-2] if len(parts) >= 2 else parts[0]
        )
        keywords = [token.lower() for token in re.findall(r"[a-z0-9]+", fine.lower()) if len(token) >= 3]
        leaves.append(
            {
                "leaf_id": re.sub(r"[^a-z0-9]+", "_", fine.lower()).strip("_")[:80],
                "classpath": classpath,
                "dept": dept,
                "class": class_name,
                "fine": fine,
                "product_name": fine,
                "template_id": _template_id_for(classpath, fine),
                "keywords": list(dict.fromkeys(keywords + [part.lower() for part in parts if len(part) >= 3])),
                "patterns": [],
            }
        )
    if not leaves:
        return None
    out = OFFICIAL_LEAVES_PATH
    try:
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps({"source_file": path.name, "leaves": leaves}, indent=1), encoding="utf-8")
    except OSError:
        out = Path("/tmp/unilog/official_leaves.json")
        out.write_text(json.dumps({"source_file": path.name, "leaves": leaves}, indent=1), encoding="utf-8")
        os.environ["UNILOG_OFFICIAL_LEAVES"] = str(out)
    print(f"Taxonomy: {len(leaves)} leaves from {path.name}")
    return out


def import_from_dir(src: Path) -> dict[str, Path | None]:
    return {
        "lov": import_lov(src),
        "uom": import_uom(src),
        "manufacturers": import_manufacturers(src),
        "fractions": import_fractions(src),
        "reference200": import_reference200(src),
        "taxonomy": import_taxonomy(src),
    }


def reset_ensure_for_tests() -> None:
    global _ENSURED
    _ENSURED = False
    _KIND_CACHE.clear()


def ensure_official_references() -> dict[str, Path | None]:
    """If a judge dropped organizer workbooks, convert them before enrich.

    Looks in guidelines/references/, guidelines/, UNILOG_REFERENCES_DIR, and
    /tmp/unilog/references. Missing files stay on mined lookups. Official
    workbooks overlay mined JSON.
    """
    global _ENSURED
    with _ENSURE_LOCK:
        if _ENSURED:
            return {}
        _ENSURED = True
        found: dict[str, Path | None] = {}
        for src in reference_search_dirs():
            if not src.exists():
                continue
            for key, path in import_from_dir(src).items():
                if path:
                    found[key] = path
        if found:
            refresh_reference_consumers(_writable_out())
        return found


def main() -> None:
    parser = argparse.ArgumentParser(description="Import dashboard reference files")
    parser.add_argument("--src", default=str(REF_DIR))
    args = parser.parse_args()
    src = Path(args.src)
    if not src.exists():
        print(f"No references directory at {src}; nothing to do.")
        print("Download the dashboard resource files into guidelines/references/ and rerun.")
        return
    results = import_from_dir(src)
    found = sum(1 for v in results.values() if v)
    print(f"\n{found}/{len(results)} reference sets imported.")
    missing = [k for k, v in results.items() if v is None]
    if missing:
        print("Missing inputs for:", ", ".join(missing))
    if found:
        refresh_reference_consumers(_writable_out())


if __name__ == "__main__":
    main()
